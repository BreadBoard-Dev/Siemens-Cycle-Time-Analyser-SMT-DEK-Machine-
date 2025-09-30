import sys
import os
import glob
import re
import datetime
import shutil
import subprocess
import platform
import time
from pathlib import Path
from statistics import mean

# PyQt6 imports
from PyQt6.QtCore import Qt, QObject, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QIcon, QPixmap
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QProgressBar, QMessageBox,
    QComboBox, QFrame, QInputDialog, QLineEdit, QDialog, QListWidget,
    QTableWidget, QTableWidgetItem, QGraphicsDropShadowEffect
)

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


# ---------------------------------------------------------
# Resource Path Helper (works for dev + PyInstaller bundle)
# ---------------------------------------------------------
def resource_path(relative_name: str):
    """
    Get absolute path to resource, works for dev and for PyInstaller exe.
    """
    if getattr(sys, 'frozen', False):  # running inside .exe
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).parent
    return base_path / relative_name


# ----------------- CONFIG -----------------
DOWNTIME_THRESHOLD = 300  # seconds (5 minutes)
REPORT_BASE_DIR = Path.home() / "Documents" / "CycleTimeReports"
MASTER_PASSWORD = "SiemensMaster2025!"
LINES_FILE = resource_path("lines.txt")
PASSWORD_FILE = resource_path("password.txt")
TEMPLATE_DB = resource_path("template.accdb")
INSTALLER_X64 = resource_path("AccessDatabaseEngine_x64.exe")
INSTALLER_X86 = resource_path("AccessDatabaseEngine_x86.exe")
# ------------------------------------------


# ---------------- ODBC Installer Worker (QThread-safe) -----------------
class ODBCInstallerWorker(QObject):
    progress = pyqtSignal(int, str)   # percent (int), message (str)
    finished = pyqtSignal(bool, str)  # success (bool), message (str)

    def __init__(self, installer_path: Path, label: str):
        super().__init__()
        self.installer_path = Path(installer_path)
        self.label = label

    def run(self):
        """Run installer in worker thread, emit progress/finished signals."""
        try:
            self.progress.emit(5, f"Starting ODBC installer ({self.label})...")

            inst_str = str(self.installer_path)
            if inst_str.lower().endswith(".msi"):
                cmd = ["msiexec", "/i", inst_str, "/quiet", "/norestart"]
            else:
                cmd = [inst_str, "/quiet", "/norestart"]

            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

            pct = 10
            while proc.poll() is None:
                time.sleep(1.5)
                pct = min(pct + 5, 95)
                self.progress.emit(pct, f"Installing ODBC driver... {pct}%")

            out, err = proc.communicate()
            try:
                with open(resource_path("odbc_install.log"), "wb") as fh:
                    fh.write(out or b"")
                    fh.write(b"\n\n----- stderr -----\n")
                    fh.write(err or b"")
            except Exception:
                pass

            if proc.returncode == 0:
                if check_odbc_driver():
                    self.progress.emit(100, "✅ ODBC driver installed.")
                    self.finished.emit(True, f"✅ ODBC driver ({self.label}) installed.")
                    return
                else:
                    if not inst_str.lower().endswith(".msi"):
                        try:
                            alt_cmd = [inst_str, "/passive", "/norestart"]
                            proc2 = subprocess.Popen(alt_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            while proc2.poll() is None:
                                time.sleep(1.0)
                                self.progress.emit(90, "Completing installation (passive)...")
                            out2, err2 = proc2.communicate()
                            with open(resource_path("odbc_install.log"), "ab") as fh:
                                fh.write(b"\n\n----- passive stderr -----\n")
                                fh.write(err2 or b"")
                        except Exception:
                            pass
                    self.finished.emit(False, "Installer finished but driver not detected. Reboot may be required.")
                    return
            else:
                if "/quiet" in " ".join(cmd) and not inst_str.lower().endswith(".msi"):
                    try:
                        alt_cmd = [inst_str, "/passive", "/norestart"]
                        proc2 = subprocess.Popen(alt_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        while proc2.poll() is None:
                            time.sleep(1.0)
                            self.progress.emit(80, "Retrying with passive mode...")
                        out2, err2 = proc2.communicate()
                        with open(resource_path("odbc_install.log"), "ab") as fh:
                            fh.write(b"\n\n----- passive stdout/stderr -----\n")
                            fh.write(out2 or b"")
                            fh.write(err2 or b"")
                        if proc2.returncode == 0 and check_odbc_driver():
                            self.progress.emit(100, "✅ ODBC driver installed (passive).")
                            self.finished.emit(True, f"✅ ODBC driver ({self.label}) installed (passive).")
                            return
                    except Exception:
                        pass

                msg = f"Installer failed: return code {proc.returncode}"
                self.finished.emit(False, msg)
                return

        except Exception as e:
            self.finished.emit(False, f"Error running installer: {e}")



# ---------------- Utility functions ----------------
def check_odbc_driver() -> bool:
    try:
        import pyodbc
        drivers = [d for d in pyodbc.drivers()]
        return any("Access" in d for d in drivers)
    except Exception:
        return False


def format_time(seconds):
    if seconds is None:
        return None
    s = int(round(seconds))
    h, r = divmod(s, 3600)
    m, s = divmod(r, 60)
    return f"{h:02}:{m:02}:{s:02}"


# ---------------- Save functions (Path-safe) ----------------
def save_to_excel(rows, path, update_progress=None):
    path = Path(path)
    # ensure parent exists
    if not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cycle_Summary"
    headers = ["Stencil", "Total_Boards", "Actual_Cycle",
               "Min_Cycle", "Max_Cycle", "Avg_Cycle", "Max_Downtime"]
    ws.append(headers)

    for r in rows:
        ws.append(r)

    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(str(path))
    if update_progress:
        update_progress(100, f"Saved Excel: Click on Open reports Button")


def save_to_access(rows, path, update_progress=None):
    path = Path(path)
    # ensure parent exists
    if not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)

    try:
        import pyodbc
    except Exception:
        raise RuntimeError("pyodbc is required for Access operations. Install ODBC & pyodbc.")

    # copy template if doesn't exist
    if not path.exists():
        if not TEMPLATE_DB.exists():
            raise FileNotFoundError("template.accdb missing in application folder.")
        shutil.copy(TEMPLATE_DB, path)

    conn_str = (r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
                f"DBQ={str(path)};")
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # ensure table exists
    try:
        cursor.execute("SELECT TOP 1 * FROM Cycle_Time")
    except Exception:
        cursor.execute("""
            CREATE TABLE Cycle_Time (
                [Stencil] TEXT,
                [Total_Boards] INT,
                [Actual_Cycle] TEXT,
                [Min_Cycle] TEXT,
                [Max_Cycle] TEXT,
                [Avg_Cycle] TEXT,
                [Max_Downtime] TEXT
            )
        """)
        conn.commit()

    for r in rows:
        cursor.execute("""
            INSERT INTO Cycle_Time
            (Stencil, Total_Boards, Actual_Cycle, Min_Cycle,
             Max_Cycle, Avg_Cycle, Max_Downtime)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, r)
    conn.commit()
    cursor.close()
    conn.close()

    if update_progress:
        update_progress(100, f"Saved Access DB: Click on Open reports Button")


# ---------------- Log parsing ----------------
def process_logs(files, update_progress=None):
    data = {}
    total = max(1, len(files))
    done = 0
    for f in files:
        done += 1
        if update_progress:
            update_progress(int(done/total*100), f"Parsing {os.path.basename(f)}...")
        try:
            with open(f, "r", errors="ignore") as fh:
                prev_dt = None
                stencil = None
                for line in fh:
                    parts = re.split(r"\s+", line.strip())
                    if len(parts) < 2:
                        continue
                    try:
                        dt = datetime.datetime.strptime(f"{parts[0]} {parts[1].split('.')[0]}", "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        continue

                    if "Product Loaded:" in line:
                        stencil = line.split("Product Loaded:")[-1].strip()
                        prev_dt = None
                        if stencil not in data:
                            data[stencil] = {"count": 0, "cycles": [], "downs": []}
                    elif "Printing board" in line and stencil:
                        data[stencil]["count"] += 1
                        if prev_dt:
                            delta = (dt - prev_dt).total_seconds()
                            if delta > 0:
                                if delta > DOWNTIME_THRESHOLD:
                                    data[stencil]["downs"].append(delta)
                                else:
                                    data[stencil]["cycles"].append(delta)
                        prev_dt = dt
        except Exception:
            continue

    rows = []
    for stencil, d in data.items():
        if d["count"] <= 1 or not d["cycles"]:
            continue
        rows.append((
            stencil,
            d["count"],
            format_time(d["cycles"][-1]),         # Actual (last)
            format_time(min(d["cycles"])),        # Min
            format_time(max(d["cycles"])),        # Max
            format_time(mean(d["cycles"])),       # Avg
            format_time(max(d["downs"])) if d["downs"] else None
        ))
    return rows


# ----------------- Admin Panel -----------------
class AdminPanel(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("🔒 Admin Panel")
        self.setModal(True)
        self.setGeometry(320, 220, 520, 520)
        self.parent_window = parent

        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("📋 Manage SMT Lines:"))
        self.line_list = QListWidget()
        self.load_lines()
        layout.addWidget(self.line_list)

        btns = QHBoxLayout()
        self.btn_add = QPushButton("➕ Add")
        self.btn_edit = QPushButton("✏️ Edit")
        self.btn_del = QPushButton("❌ Delete")
        btns.addWidget(self.btn_add); btns.addWidget(self.btn_edit); btns.addWidget(self.btn_del)
        layout.addLayout(btns)

        layout.addSpacing(8)
        self.btn_change_pwd = QPushButton("🔑 Change Password")
        self.btn_install_odbc = QPushButton("💾 Install ODBC Driver")
        layout.addWidget(self.btn_change_pwd)
        layout.addWidget(self.btn_install_odbc)

        # connects
        self.btn_add.clicked.connect(self.add_line)
        self.btn_edit.clicked.connect(self.edit_line)
        self.btn_del.clicked.connect(self.delete_line)
        self.btn_change_pwd.clicked.connect(self.change_password)
        self.btn_install_odbc.clicked.connect(self.install_odbc_driver)

        # holder for installer thread/worker
        self._installer_thread = None
        self._installer_worker = None

    def load_lines(self):
        self.line_list.clear()
        if not LINES_FILE.exists():
            with open(LINES_FILE, "w", encoding="utf-8") as f:
                f.write("\n".join(["SMT Line 1", "SMT Line 2", "SMT Line 3", "SMT Line 4"]))
        with open(LINES_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    self.line_list.addItem(line.strip())

    def save_lines(self):
        with open(LINES_FILE, "w", encoding="utf-8") as f:
            for i in range(self.line_list.count()):
                f.write(self.line_list.item(i).text() + "\n")

    def add_line(self):
        text, ok = QInputDialog.getText(self, "Add SMT Line", "Enter new line name:")
        if not ok or not text.strip():
            return
        line_name = text.strip()
        self.line_list.addItem(line_name)
        self.save_lines()

        # Create folders and copy template for Access
        try:
            base = REPORT_BASE_DIR / line_name.replace(" ", "_")
            (base / "ExcelReports").mkdir(parents=True, exist_ok=True)
            access_dir = base / "AccessReports"
            access_dir.mkdir(parents=True, exist_ok=True)
            accdb_path = access_dir / f"Line_{line_name.replace(' ', '_')}.accdb"
            if not accdb_path.exists():
                if TEMPLATE_DB.exists():
                    shutil.copy(TEMPLATE_DB, accdb_path)
                    QMessageBox.information(self, "Access DB Created",
                                            f"✅ Access DB created for {line_name} at:\n{accdb_path}")
                else:
                    QMessageBox.warning(self, "Template Missing",
                                        "⚠ template.accdb not found in app folder; cannot create Access DB.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not create report folders / DB:\n{e}")

    def edit_line(self):
        current = self.line_list.currentItem()
        if not current:
            QMessageBox.information(self, "Select", "Select a line to edit.")
            return
        old_name = current.text()
        text, ok = QInputDialog.getText(self, "Edit SMT Line", "Rename line:", text=old_name)
        if not ok or not text.strip():
            return
        new_name = text.strip()
        current.setText(new_name)
        self.save_lines()
        # rename folder if exists
        try:
            old_dir = REPORT_BASE_DIR / old_name.replace(" ", "_")
            new_dir = REPORT_BASE_DIR / new_name.replace(" ", "_")
            if old_dir.exists() and not new_dir.exists():
                old_dir.rename(new_dir)
        except Exception as e:
            QMessageBox.warning(self, "Warning", f"Renaming folder failed:\n{e}")

    def delete_line(self):
        row = self.line_list.currentRow()
        if row < 0:
            return
        line_name = self.line_list.item(row).text()
        confirm = QMessageBox.question(self, "Confirm Delete",
                                       f"Delete SMT line '{line_name}' and its report folder?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if confirm != QMessageBox.StandardButton.Yes:
            return
        self.line_list.takeItem(row)
        self.save_lines()
        try:
            reports_dir = REPORT_BASE_DIR / line_name.replace(" ", "_")
            if reports_dir.exists():
                shutil.rmtree(reports_dir)
                QMessageBox.information(self, "Deleted", f"✅ Removed folder {reports_dir}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not delete reports folder:\n{e}")

    def change_password(self):
        new_pass, ok = QInputDialog.getText(self, "Change Password", "Enter new password:", QLineEdit.EchoMode.Password)
        if not ok or not new_pass.strip():
            return
        try:
            with open(PASSWORD_FILE, "w", encoding="utf-8") as f:
                f.write(new_pass.strip())
            QMessageBox.information(self, "Saved", "✅ New admin password saved.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not save password:\n{e}")

    def install_odbc_driver(self):
        """
        Detect Office bitness, choose appropriate installer (x64/x86),
        then run worker inside QThread to avoid GUI thread updates from background.
        """
        import winreg

        # detect office bitness (fallback to system arch)
        office_bitness = None
        try:
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                 r"SOFTWARE\Microsoft\Office\ClickToRun\Configuration")
            val, _ = winreg.QueryValueEx(key, "Platform")
            if "x64" in val.lower():
                office_bitness = "x64"
            elif "x86" in val.lower():
                office_bitness = "x86"
        except Exception:
            arch, _ = platform.architecture()
            office_bitness = "x64" if "64" in arch else "x86"

        preferred = INSTALLER_X64 if office_bitness == "x64" else INSTALLER_X86
        fallback = INSTALLER_X86 if office_bitness == "x64" else INSTALLER_X64

        candidates = []
        if preferred.exists():
            candidates.append((preferred, office_bitness))
        if fallback.exists():
            candidates.append((fallback, "fallback"))

        if not candidates:
            QMessageBox.warning(self, "Installer Missing",
                                f"Place {INSTALLER_X64.name} or {INSTALLER_X86.name} in the app folder.")
            return

        # Start a worker thread for the first candidate
        installer_path, label = candidates[0]

        # Create worker and thread
        self._installer_thread = QThread()
        self._installer_worker = ODBCInstallerWorker(installer_path, label)
        self._installer_worker.moveToThread(self._installer_thread)

        # connect signals
        self._installer_thread.started.connect(self._installer_worker.run)
        self._installer_worker.progress.connect(self.parent_window.update_progress)
        self._installer_worker.finished.connect(self.on_install_finished)

        # cleanup connections
        self._installer_worker.finished.connect(self._installer_thread.quit)
        self._installer_worker.finished.connect(self._installer_worker.deleteLater)
        self._installer_thread.finished.connect(self._installer_thread.deleteLater)

        # start thread
        self._installer_thread.start()

    def on_install_finished(self, success: bool, message: str):
        # This runs in main thread because finished is a Qt signal
        if success:
            QMessageBox.information(self, "ODBC Install", message)
        else:
            # If first attempt failed, offer option to try the other installer (fallback)
            reply = QMessageBox.question(self, "Install Failed",
                                         f"{message}\n\nTry fallback installer if available?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                # find fallback installer and run it
                arch, _ = platform.architecture()
                prefer_x64 = "64" in arch
                if prefer_x64:
                    primary, fallback = INSTALLER_X64, INSTALLER_X86
                else:
                    primary, fallback = INSTALLER_X86, INSTALLER_X64
                if fallback.exists():
                    # start fallback install
                    self._installer_thread = QThread()
                    self._installer_worker = ODBCInstallerWorker(fallback, "fallback")
                    self._installer_worker.moveToThread(self._installer_thread)
                    self._installer_thread.started.connect(self._installer_worker.run)
                    self._installer_worker.progress.connect(self.parent_window.update_progress)
                    self._installer_worker.finished.connect(self.on_install_finished)
                    self._installer_worker.finished.connect(self._installer_thread.quit)
                    self._installer_worker.finished.connect(self._installer_worker.deleteLater)
                    self._installer_thread.finished.connect(self._installer_thread.deleteLater)
                    self._installer_thread.start()
                else:
                    QMessageBox.warning(self, "Fallback Missing", f"Fallback installer not found: {fallback}")
            else:
                QMessageBox.warning(self, "Install", "ODBC installation did not complete successfully.")

        # refresh parent's ODBC UI status
        try:
            self.parent_window.refresh_odbc_ui()
        except Exception:
            pass


# ----------------- UI Panels -----------------
class FuturisticPanel(QFrame):
    def __init__(self, title, widget):
        super().__init__()
        self.setObjectName("FuturisticPanel")
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(18)
        shadow.setOffset(0, 2)
        shadow.setColor(QColor(0, 0, 0, 60))
        self.setGraphicsEffect(shadow)
        layout = QVBoxLayout(self)
        lbl = QLabel(title)
        lbl.setObjectName("PanelTitle")
        layout.addWidget(lbl)
        layout.addWidget(widget)


# ----------------- Main UI -----------------
class CycleAnalyzerUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Siemens Cycle Time Analyzer")

        icon_path = resource_path("diagram.ico")
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        self.setGeometry(160, 180, 1200, 720)

        self.files = []
        self.save_path = None

        central = QWidget()
        self.setCentralWidget(central)
        main = QHBoxLayout(central)

        # Left controls
        left = QVBoxLayout()
        left.setSpacing(16)

        # File selection
        fw = QWidget()
        fl = QHBoxLayout(fw)
        self.btn_files = QPushButton("📂 Select Files")
        self.btn_folder = QPushButton("📁 Select Folder")
        self.lbl_files = QLabel("No files selected")
        self.lbl_files.setWordWrap(True)
        fl.addWidget(self.btn_files); fl.addWidget(self.btn_folder); fl.addWidget(self.lbl_files)
        left.addWidget(FuturisticPanel("Input Selection", fw))

        # Output settings
        sw = QWidget()
        sl = QHBoxLayout(sw)
        self.cmb_save = QComboBox()
        self.cmb_save.addItems(["Excel", "Access"])
        self.cmb_line = QComboBox()
        self.load_lines()
        sl.addWidget(QLabel("Save As:")); sl.addWidget(self.cmb_save)
        sl.addWidget(QLabel("SMT Line:")); sl.addWidget(self.cmb_line)
        left.addWidget(FuturisticPanel("Output Settings", sw))

        # Execution buttons
        rw = QWidget()
        rl = QVBoxLayout(rw)
        self.btn_run = QPushButton("⏳ Run Analysis")
        self.btn_open = QPushButton("📁 Open Reports Folder")
        self.btn_load = QPushButton("📄 Load Report")
        self.btn_admin = QPushButton("🔒 Admin Panel")
        rl.addWidget(self.btn_run)
        rl.addWidget(self.btn_open)
        rl.addWidget(self.btn_load)
        rl.addWidget(self.btn_admin)
        left.addWidget(FuturisticPanel("Execution", rw))

        # Status panel
        stw = QWidget()
        stl = QVBoxLayout(stw)
        self.progress = QProgressBar()
        self.lbl_status = QLabel("Ready")
        stl.addWidget(self.progress); stl.addWidget(self.lbl_status)
        left.addWidget(FuturisticPanel("Status", stw))
        left.addStretch()

        # Right table
        right = QVBoxLayout()
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["Stencil", "Total_Boards", "Actual_Cycle", "Min_Cycle", "Max_Cycle", "Avg_Cycle", "Max_Downtime"]
        )
        self.table.horizontalHeader().setStretchLastSection(True)
        right.addWidget(FuturisticPanel("Cycle Time Summary", self.table))

        # Logo + credits
        logo_label = QLabel()
        logo_path = resource_path("siemens.png")
        if logo_path.exists():
            pixmap = QPixmap(str(logo_path))
            logo_label.setPixmap(pixmap.scaled(200, 60, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

        credit_label = QLabel("Developed by <b>Prasad Gawas</b>")
        credit_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        credit_label.setStyleSheet("color: #1976d2; font-size: 10pt; font-family: Segoe UI;")

        footer = QVBoxLayout()
        footer.addWidget(logo_label)
        footer.addWidget(credit_label)
        footer_widget = QWidget()
        footer_widget.setLayout(footer)
        right.addWidget(footer_widget, alignment=Qt.AlignmentFlag.AlignCenter)

        main.addLayout(left, 1)
        main.addLayout(right, 2)

        self.apply_white_theme()

        # Connections
        self.btn_files.clicked.connect(self.select_files)
        self.btn_folder.clicked.connect(self.select_folder)
        self.btn_run.clicked.connect(self.run_analysis)
        self.btn_open.clicked.connect(self.open_reports)
        self.btn_load.clicked.connect(self.load_report)
        self.btn_admin.clicked.connect(self.open_admin)

        # initial odbc status -> update UI
        self.refresh_odbc_ui()

    def apply_white_theme(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #F9FBFD; }
            QLabel { color: #222; font-size: 11pt; font-family: 'Segoe UI'; }
            QLabel#PanelTitle { font-size: 13pt; font-weight: bold; color: #1976d2; }
            QPushButton {
                background-color: rgba(255,255,255,0.95);
                color: #1976d2;
                border: 2px solid #1976d2;
                border-radius: 10px;
                padding: 8px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1976d2; color: white; }
            QComboBox {
                background-color: rgba(255,255,255,0.95);
                color: #1976d2;
                border: 2px solid #FF4081;
                border-radius: 8px; padding: 6px; font-weight: bold;
            }
            QProgressBar { border: 2px solid #bbb; border-radius: 10px;
                text-align: center; font-weight: bold; background-color: #f0f0f0; }
            QProgressBar::chunk { background-color: #76FF03; border-radius: 8px; }
            QFrame#FuturisticPanel { background-color: rgba(255,255,255,0.88);
                border: 1px solid rgba(25,118,210,0.12); border-radius: 12px; padding: 10px; }
        """)

    # --- Lines ---
    def load_lines(self):
        if not LINES_FILE.exists():
            with open(LINES_FILE, "w", encoding="utf-8") as f:
                f.write("\n".join(["SMT Line 1", "SMT Line 2", "SMT Line 3", "SMT Line 4"]))
        self.cmb_line.clear()
        with open(LINES_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    self.cmb_line.addItem(line.strip())

    # --- ODBC UI ---
    def refresh_odbc_ui(self):
        installed = check_odbc_driver()
        idx = self.cmb_save.findText("Access")

        if installed:
            # If "Access" isn't present, add it; otherwise enable its item safely.
            if idx == -1:
                self.cmb_save.addItem("Access")
            else:
                model = self.cmb_save.model()
                if model is not None:
                    combo_item = model.item(idx)
                    if combo_item is not None:
                        combo_item.setEnabled(True)
            self.lbl_status.setText("✅ ODBC driver installed.")
        else:
            # If present, disable its item safely.
            if idx != -1:
                model = self.cmb_save.model()
                if model is not None:
                    combo_item = model.item(idx)
                    if combo_item is not None:
                        combo_item.setEnabled(False)
            self.lbl_status.setText("⚠ ODBC driver missing — Access disabled.")

    # --- Results display ---
    def show_results(self, rows):
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self.table.setItem(r, c, QTableWidgetItem(str(val if val is not None else "")))

    # --- File selection ---
    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Log Files", "", "Log/Text Files (*.txt *.log)")
        if files:
            self.files = list(files)
            display = ", ".join(os.path.basename(f) for f in files)
            if len(display) > 180:
                display = f"{len(files)} files selected"
            self.lbl_files.setText(display)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder Containing Log Files")
        if folder:
            self.files = glob.glob(os.path.join(folder, "*.txt")) + glob.glob(os.path.join(folder, "*.log"))
            self.lbl_files.setText(f"{len(self.files)} files selected")

    # --- Run analysis & save ---
    def run_analysis(self):
        if not self.files:
            QMessageBox.warning(self, "No Files", "Please select logs.")
            return

        line_name = self.cmb_line.currentText().strip()
        if not line_name:
            QMessageBox.warning(self, "No Line", "Select SMT line.")
            return

        base_dir = REPORT_BASE_DIR / line_name.replace(" ", "_")
        if self.cmb_save.currentText() == "Excel":
            out_dir = base_dir / "ExcelReports"
            out_dir.mkdir(parents=True, exist_ok=True)
            save_path = out_dir / f"Summary_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx"
        else:
            out_dir = base_dir / "AccessReports"
            out_dir.mkdir(parents=True, exist_ok=True)
            save_path = out_dir / f"Line_{line_name.replace(' ', '_')}.accdb"

        rows = process_logs(self.files, self.update_progress)
        if not rows:
            QMessageBox.information(self, "No Data", "No valid cycle times found.")
            return

        try:
            if self.cmb_save.currentText() == "Excel":
                save_to_excel(rows, save_path, self.update_progress)
            else:
                # Ensure pyodbc available
                try:
                    import pyodbc  # noqa: F401
                except Exception:
                    QMessageBox.critical(self, "pyodbc Missing", "pyodbc is required. Install ODBC & pyodbc.")
                    return
                save_to_access(rows, save_path, self.update_progress)
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", f"Failed to save report:\n{e}")
            return

        self.show_results(rows)
        QMessageBox.information(self, "Done", f"Saved → {save_path}")

    # --- Open folder/report ---
    def open_reports(self):
        line_name = self.cmb_line.currentText().strip()
        if not line_name:
            QMessageBox.warning(self, "No Line", "Select SMT line.")
            return
        path = REPORT_BASE_DIR / line_name.replace(" ", "_")
        path.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(str(path))
        except Exception as e:
            QMessageBox.warning(self, "Open Failed", f"Could not open folder:\n{e}")

    def load_report(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Report File",
            str(REPORT_BASE_DIR),
            "Reports (*.xlsx *.accdb);;Excel (*.xlsx);;Access (*.accdb)"
        )
        if not file_path:
            return

        file_path = Path(file_path)
        rows = []
        if file_path.suffix.lower() == ".xlsx":
            wb = load_workbook(str(file_path))
            ws = wb.active
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx == 0:
                    continue
                rows.append(tuple(row))
        elif file_path.suffix.lower() == ".accdb":
            try:
                import pyodbc
            except Exception:
                QMessageBox.critical(self, "pyodbc Missing", "pyodbc is required to read Access files.")
                return
            conn_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={str(file_path)};"
            try:
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM Cycle_Time")
                for row in cursor.fetchall():
                    rows.append(tuple(row))
                cursor.close()
                conn.close()
            except Exception as e:
                QMessageBox.critical(self, "Read Failed", f"Failed to read Access DB:\n{e}")
                return
        else:
            QMessageBox.information(self, "Unsupported", "Unsupported file type.")
            return

        if not rows:
            QMessageBox.information(self, "Empty", "No data found in report.")
            return

        self.show_results(rows)
        QMessageBox.information(self, "Report Loaded", f"✅ Loaded report:\n{file_path}")

    # --- Admin access ---
    def open_admin(self):
        stored = MASTER_PASSWORD
        if PASSWORD_FILE.exists():
            try:
                with open(PASSWORD_FILE, "r", encoding="utf-8") as f:
                    val = f.read().strip()
                    if val:
                        stored = val
            except Exception:
                pass

        entered, ok = QInputDialog.getText(self, "Admin Access", "Enter admin password (or master):", QLineEdit.EchoMode.Password)
        if not ok:
            return
        if entered.strip() != stored and entered.strip() != MASTER_PASSWORD:
            QMessageBox.critical(self, "Access Denied", "❌ Incorrect password.")
            return

        dlg = AdminPanel(self)
        dlg.exec()
        # After admin actions, reload lines and refresh ODBC UI
        self.load_lines()
        # refresh ODBC UI
        self.refresh_odbc_ui()

    # --- Progress helper ---
    def update_progress(self, val, msg):
        try:
            self.progress.setValue(int(val))
        except Exception:
            pass
        self.lbl_status.setText(str(msg))
        QApplication.processEvents()


# ----------------- Main -----------------
def main():
    app = QApplication(sys.argv)
    icon_path = resource_path("diagram.ico")
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))
    ui = CycleAnalyzerUI()
    ui.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
